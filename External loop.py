# -*- coding: utf-8 -*-
"""
Created on Tue Dec  3 11:04:23 2019

@author: Alessandro Barbieri
            Politecnico di Milano - Department of Energy
"""
import openpyxl
import calliope
import os
#import matplotlib.pyplot as plt


#calliope.set_log_level('INFO')


for indice in range(0,5):                 
    
      model = calliope.Model('model.yaml')

      model.run()      
      
      #estraggo dei dati giusto per avere la stessa struttura (pandas Daraframe) anche per le variabili che andrò a creare
      DataFrame1=model.get_formatted_array('carrier_prod').loc[{'techs':'PV','carriers':'power','locs':['Zambia']}].to_pandas().T #estraggo per usarlo come base per costruire variabile salto nuovo
      DataFrame2=model.get_formatted_array('carrier_prod').loc[{'techs':'PV','carriers':'power','locs':['Zambia']}].to_pandas().T
      DataFrame3=model.get_formatted_array('carrier_prod').loc[{'techs':'PV','carriers':'power','locs':['Zambia']}].to_pandas().T
      DataFrame4=model.get_formatted_array('carrier_prod').loc[{'techs':'PV','carriers':'power','locs':['Zambia']}].to_pandas().T
      DataFrame5=model.get_formatted_array('carrier_prod').loc[{'techs':'PV','carriers':'power','locs':['Zambia']}].to_pandas().T
      DataFrame6=model.get_formatted_array('carrier_prod').loc[{'techs':'PV','carriers':'power','locs':['Zambia']}].to_pandas().T
      DataFrame7=model.get_formatted_array('carrier_prod').loc[{'techs':'PV','carriers':'power','locs':['Zambia']}].to_pandas().T
      DataFrame8=model.get_formatted_array('carrier_prod').loc[{'techs':'PV','carriers':'power','locs':['Zambia']}].to_pandas().T
      DataFrame9=model.get_formatted_array('carrier_prod').loc[{'techs':'PV','carriers':'power','locs':['Zambia']}].to_pandas().T
      DataFrame10=model.get_formatted_array('carrier_prod').loc[{'techs':'PV','carriers':'power','locs':['Zambia']}].to_pandas().T
      DataFrame11=model.get_formatted_array('carrier_prod').loc[{'techs':'PV','carriers':'power','locs':['Zambia']}].to_pandas().T
      DataFrame12=model.get_formatted_array('carrier_prod').loc[{'techs':'PV','carriers':'power','locs':['Zambia']}].to_pandas().T
      DataFrame13=model.get_formatted_array('carrier_prod').loc[{'techs':'OCGT','carriers':'power','locs':['Moz-North-Center']}].to_pandas().T
      DataFrame14=model.get_formatted_array('carrier_prod').loc[{'techs':'OCGT','carriers':'power','locs':['Moz-North-Center']}].to_pandas().T
      DataFrame15=model.get_formatted_array('carrier_prod').loc[{'techs':'OCGT','carriers':'power','locs':['Moz-North-Center']}].to_pandas().T
      
      #ITT    
      #estraggo valore dello storage e inizializzo le variabili eff, water to storage a valle, superficie del baciono e perdite da evaporazione
      #inoltre carico file excell in cui sono riportati i valori dell'evaporation rate
      StorageA_nuovo=model.get_formatted_array('storage').loc[{'techs': 'storageA','locs':['Zambia']}].to_pandas().T #estraggo risultati da modello
      eff_conv_ITT = DataFrame1
      WaterToStorageB_ITT = DataFrame2
      evapLoss_ITT= DataFrame3
      supStorage_ITT=DataFrame4
      Data_ITT=openpyxl.load_workbook('Timeseries/lsv_min_max_rel_ITT.xlsx') #carico excel con tutti i dati relativi a ITT
      EvapRate_ITT=Data_ITT['evap+salto'] #seleziono il foglio dove ci sono i dati di evaporazione, i quali si trovano nella sesta colonna, la selezioneremo successivamente

      #ciclo per calcolare in ogni timestep il valore di eff e delle altre variabili inizializzate
      for a in range(0, 48):    
          eff_conv_ITT.iloc[a]=(40.5-(1030.5-(-5*10**(-19)*(StorageA_nuovo.iloc[a] + 699000000)**2 + 8*10**(-9)*(StorageA_nuovo.iloc[a] + 699000000)+1000.7)))*9.8*1000*0.9/3600*10**(-3)
          WaterToStorageB_ITT.iloc[a]=eff_conv_ITT.iloc[a]**(-1)
          supStorage_ITT.iloc[a]= -3*10**(-12)*(StorageA_nuovo.iloc[a] + 699000000)**2 + 0.08*(StorageA_nuovo.iloc[a] + 699000000)+3*10**7
          b=a+2
          evapLoss_ITT.iloc[a]= (supStorage_ITT.iloc[a]*EvapRate_ITT.cell(row=b, column=6).value/1000/EvapRate_ITT.cell(row=b, column=2).value/24)/(StorageA_nuovo.iloc[a] + 699000000)
          
      #rimuovi le timeseries presenti nella directory e le sostituisco con quelle aggiornate    
      os.remove('Timeseries/effITT.txt')    
      os.remove('Timeseries/WaterToStorageB_ITT.txt')
      os.remove('Timeseries/evapLoss_ITT.txt')
      eff_conv_ITT.to_csv('Timeseries/effITT.txt') 
      WaterToStorageB_ITT.to_csv('Timeseries/WaterToStorageB_ITT.txt')
      evapLoss_ITT.to_csv('Timeseries/evapLoss_ITT.txt')

      #Kafue Gorge
      StorageB_nuovo=model.get_formatted_array('storage').loc[{'techs': 'storageB','locs':['Zambia']}].to_pandas().T #estraggo risultati da modello
      eff_conv_KG = DataFrame5
      WaterToStorageD_KG = DataFrame6
      evapLoss_KG= DataFrame7
      supStorage_KG=DataFrame8
      Data_KG=openpyxl.load_workbook('Timeseries/lsv_min_max_rel_KGU.xlsx') #carico excel con tutti i dati relativi a ITT
      EvapRate_KG=Data_KG['evap+salto'] #seleziono il foglio dove ci sono i dati di evaporazione, i quali si trovano nella sesta colonna, la selezioneremo successivamente

      for k in range(0, 48):    
          eff_conv_KG.iloc[k]=(397-(977.6-(957*(StorageB_nuovo.iloc[k] + 5000000)**0.001)))*9.8*1000*0.9/3600*10**(-3)
          WaterToStorageD_KG.iloc[k]=eff_conv_KG.iloc[k]**(-1)
          supStorage_KG.iloc[k]= -10**(-10)*(StorageB_nuovo.iloc[k] + 5000000)**2 + 1.129*(StorageB_nuovo.iloc[k] + 5000000)-10**7
          h=k+2
          evapLoss_KG.iloc[k]= (supStorage_KG.iloc[k]*EvapRate_KG.cell(row=h,column=6).value/1000/EvapRate_KG.cell(row=h,column=2).value/24)/(StorageB_nuovo.iloc[k] + 5000000)
                    
      os.remove('Timeseries/effKG.txt')    
      os.remove('Timeseries/WaterToStorageD_KG.txt')
      os.remove('Timeseries/evapLoss_KG.txt')
      eff_conv_KG.to_csv('Timeseries/effKG.txt') 
      WaterToStorageD_KG.to_csv('Timeseries/WaterToStorageD_KG.txt')
      evapLoss_KG.to_csv('Timeseries/evapLoss_KG.txt')
 
      #Kariba
      StorageC_nuovo=model.get_formatted_array('storage').loc[{'techs': 'storageC','locs':['Zambia']}].to_pandas().T #estraggo risultati da modello
      eff_conv_KA = DataFrame9
      WaterToStorageD_KA = DataFrame10
      evapLoss_KA= DataFrame11
      supStorage_KA=DataFrame12
      Data_KA=openpyxl.load_workbook('Timeseries/lsv_min_max_rel_KA.xlsx') #carico excel con tutti i dati relativi a ITT
      EvapRate_KA=Data_KA['evap+salto'] #seleziono il foglio dove ci sono i dati di evaporazione, i quali si trovano nella sesta colonna, la selezioneremo successivamente

      for j in range(0, 48):    
          eff_conv_KA.iloc[j]=(110-(489.5-(-5*10**(-23)*(StorageC_nuovo.iloc[j] + 116054000000)**2 + 2*10**(-10)*(StorageC_nuovo.iloc[j] + 116054000000)+452.89)))*9.8*1000*0.9/3600*10**(-3)
          WaterToStorageD_KA.iloc[j]=eff_conv_KA.iloc[j]**(-1)
          supStorage_KA.iloc[j]= -10**(-13)*(StorageC_nuovo.iloc[j] + 116054000000)**2 + 0.0493*(StorageC_nuovo.iloc[j] + 116054000000)+4*10**6
          l=j+2
          evapLoss_KA.iloc[j]= (supStorage_KA.iloc[j]*EvapRate_KA.cell(row=l,column=6).value/1000/EvapRate_KA.cell(row=l,column=2).value/24)/(StorageC_nuovo.iloc[j] + 116054000000)
              
      os.remove('Timeseries/effKA.txt')    
      os.remove('Timeseries/WaterToStorageD_KA.txt')
      os.remove('Timeseries/evapLoss_KA.txt')
      eff_conv_KA.to_csv('Timeseries/effKA.txt') 
      WaterToStorageD_KA.to_csv('Timeseries/WaterToStorageD_KA.txt')
      evapLoss_KA.to_csv('Timeseries/evapLoss_KA.txt')
 
      #Cahora
      StorageD_nuovo=model.get_formatted_array('storage').loc[{'techs': 'storageD','locs':['Moz-North-Center']}].to_pandas().T #estraggo risultati da modello
      eff_conv_CB = DataFrame13 #creo variabile eff   
      evapLoss_CB= DataFrame14
      supStorage_CB=DataFrame15
      Data_CB=openpyxl.load_workbook('Timeseries/lsv_min_max_rel_CB.xlsx') #carico excel con tutti i dati relativi a ITT
      EvapRate_CB=Data_CB['evap+salto'] #seleziono il foglio dove ci sono i dati di evaporazione, i quali si trovano nella sesta colonna, la selezioneremo successivamente


      for i in range(0, 48):    
          eff_conv_CB.iloc[i]=(128-(331-(6*10**(-21)*(StorageD_nuovo.iloc[i] + 32000000)**2 + 9*10**(-10)*(StorageD_nuovo.iloc[i] + 32000000)+294.16)))*9.8*1000*0.9/3600*10**(-3)  #riempo variabile eff con eff calcolata in funzione dello storage
          supStorage_CB.iloc[i]= -2*10**(-13)*(StorageD_nuovo.iloc[i] + 32000000)**2 + 0.0469*(StorageD_nuovo.iloc[i] + 32000000)+8*10**8
          c=i+2
          evapLoss_CB.iloc[i]= (supStorage_CB.iloc[i]*EvapRate_CB.cell(row=c,column=6).value/1000/EvapRate_CB.cell(row=c,column=2).value/24)/(StorageD_nuovo.iloc[i] + 32000000)
              
      os.remove('Timeseries/effCB.txt')
      os.remove('Timeseries/evapLoss_CB.txt')              
      eff_conv_CB.to_csv('Timeseries/effCB.txt') #produco csv con eff
      evapLoss_CB.to_csv('Timeseries/evapLoss_CB.txt')


      #plot degli storage (criterio di convergenzaq: assenza di variazioni significative nell'andamento tra un'iterazione e l'altra)
      #day = '2005-01-01 00:00:00'
      #end = '2006-12-31 23:00:00'

         
      #fig, (ax1) = plt.subplots(1, figsize=(8,6))         
      #ax1.plot(StorageA_nuovo[day:end].index,StorageA_nuovo[day:end].values,'#0071A0', alpha=0.2, label='ITT')
      #ax1.plot(StorageB_nuovo[day:end].index,StorageB_nuovo[day:end].values,'#A04000', alpha=0.2, label='KG')
      #fig.savefig("ITT-KG.png", dpi=fig.dpi,bbox_inches='tight')
               
      #fig, (ax1) = plt.subplots(1, figsize=(8,6))                 
      #ax1.plot(StorageC_nuovo[day:end].index,StorageC_nuovo[day:end].values,'#F9E900', alpha=0.2, label='KA')         
      #ax1.plot(StorageD_nuovo[day:end].index,StorageD_nuovo[day:end].values,'#0015F9', alpha=0.2, label='CB')         
      #fig.savefig("KA-CB.png", dpi=fig.dpi,bbox_inches='tight')   

model.to_csv('C:/Users/stevo/Repositories/ZRB_Calliope-Hydro/Results/Scenariox/Iterazione'+str(indice))
     

